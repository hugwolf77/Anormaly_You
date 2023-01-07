# 2022.11.22
# DLinear model .py for TTA validation Docker

import os
import argparse

import torch
from datetime import date, datetime
import time
# from dataclasses import dataclass, field

from typing import Set, List, Dict, Tuple #, final
from openpyxl import Workbook, load_workbook
# import json

from SFW_exp.exp_main import Exp_Main
from dataform import TS_DATA, args
from TS_ANLYS import *
from Toolkit import LOAD_TS_FILE, REPORT
import random
import numpy as np
# import pandas as pd

from scipy.stats import pearsonr
from sklearn.metrics import *

from pickle import FALSE

# cmd imputed option & parser
parser = argparse.ArgumentParser(description='GAIN_model for data reGeneration. version b.0.1')
# parser.add_argument('select', help ='select to "train" or "test"', default='train')
parser.add_argument('-data_path','-d', help ='load -input data - for model', default=None)
parser.add_argument('-save_path','-s', help ='save -save result - for model', default=None)
parser.add_argument('-file_name','-p', help ='save result_file_name - xlsx - for model', default=None)
arg_obtion = parser.parse_args()

fix_seed = 2022
random.seed(fix_seed)
torch.manual_seed(fix_seed)
np.random.seed(fix_seed)

class TSFAR:
    def __init__(self, file_path, save_path, file_format, meta_file_format, fxls_name, sheet_name, sheet_col) -> None:
        self.file_path = file_path
        self.save_path = save_path
        self.file_format = file_format
        self.meta_file_format = meta_file_format
        self.file_list = [f"{file}" for file in os.listdir(self.file_path+'csv/') if self.file_format in file]
        self.file_list.sort()
        self.meta_file_list = [f"{file}" for file in os.listdir(self.file_path+'json/') if self.meta_file_format in file]
        self.meta_file_list.sort()
        self.file_list_len = len(self.file_list) 
        print(f" --- Find Files for load : {self.file_list_len}\n")
        self.meta_file_list_len = len(self.meta_file_list)
        print(f" --- Find Meta Files for load : {self.meta_file_list_len}\n")

        # self.TS_DATA_ANLYS_LIST = []
        self.fxls_name = fxls_name
        self.sheet_name = sheet_name
        self.sheet_col = ['Id','Modi_Date','FileName','FilePath', 'Row_count', 'ColumnNames', 'Row_NA_count','plot']

    def TS_DATA_ANLYS_FILES(self):

        if self.fxls_name in os.listdir(self.file_path):
          wb = load_workbook(filename = os.path.join(self.save_path+self.fxls_name))
        else:
          wb = Workbook()
        write_ws = wb.create_sheet(self.sheet_name,-1)
        fig_ws = wb.create_sheet(self.sheet_name+'_plot',-1)
        write_ws = wb.active
        colname = ['순번','measure_id','reservoir','name','target_value','input_col','modify_day','file_name','file_path','row_count', 'MAE', 'MSE','RMSE','MAPE','MSPE','RSE','CORR','start-time','time-cost'] 

        colname.extend([])
        write_ws.append(colname)

        for file_ix in range(self.file_list_len):
            target_ix = file_ix # 같은 파일 순서로 전제
            print(f"\n=================================================================>> : ")
            print(f"Total Task file Count.... : {self.file_list_len}")
            print(f"Total Meta file Count.... : {self.meta_file_list_len}")
            print(f"File Processing.... Now ...... : {file_ix + 1}")
            Meta_info = LOAD_TS_FILE(self.file_path, self.file_format, self.meta_file_format, file_ix, target_ix).Meta_file_loader
            in_cols, in_cols_len, tg_cols, tg_cols_len, _ = Meta_info()
            print(f".......Total_Input value ........: {in_cols}, ..... input_cols_len : {in_cols_len}")
            for i, tg in enumerate(tg_cols):
              print(f".......TotalTarget value loop.........: {tg_cols_len}")
              print(f"............Target value loop.........: {i+1}, target : {tg}")
              now = time.localtime()
              start_time = time.strftime('%Y%m%d %I:%M:%S %p', now)
              start = time.time()
              TS_DATA = TS_ANLYS(self.file_path, self.file_format, self.meta_file_format, file_ix, target_ix, tg).Set_TS_Data()
              # self.TS_DATA_ANLYS_LIST.append(TS_DATA)
              time_cost = time.time() - start

              row_max = write_ws.max_row + 1
              col_max = write_ws.max_column # need sheet col lenth check
              write_ws.cell(row_max,1,TS_DATA.Id)
              write_ws.cell(row_max,2,TS_DATA.ob_measure_id)
              write_ws.cell(row_max,3,TS_DATA.ob_reservoir)
              write_ws.cell(row_max,4,TS_DATA.ob_name)
              write_ws.cell(row_max,5,TS_DATA.ob_target)
              write_ws.cell(row_max,6,TS_DATA.ob_input)
              write_ws.cell(row_max,7,TS_DATA.Modi_Date)
              write_ws.cell(row_max,8,TS_DATA.FileName)
              write_ws.cell(row_max,9,TS_DATA.FilePath)
              write_ws.cell(row_max,10,TS_DATA.Row_count)
              write_ws.cell(row_max,11,TS_DATA.MAE)
              write_ws.cell(row_max,12,TS_DATA.MSE)
              write_ws.cell(row_max,13,TS_DATA.RMSE)
              write_ws.cell(row_max,14,TS_DATA.MAPE)
              write_ws.cell(row_max,15,TS_DATA.MSPE)
              write_ws.cell(row_max,16,TS_DATA.RSE)
              write_ws.cell(row_max,17,str(TS_DATA.CORR))
              write_ws.cell(row_max,18,start_time)
              write_ws.cell(row_max,19,time_cost)
              wb.save(self.save_path+self.fxls_name)
              print(f"=================================================================>> : \n")
        return #TS_DATA

def main(file_path, save_path, file_name, file_format=".csv", meta_file_format = ".json",sheet_col = ""):
    print ("[--------------------DLinear Model Start------------------]")
    fxls_name =  str(file_name + ".xlsx")
    sheet_name = "SFW_DLinear_pred24hour"
    sheet_col = ""
    autoAR = TSFAR(file_path, save_path, file_format, meta_file_format, fxls_name, sheet_name, sheet_col)
    autoAR.TS_DATA_ANLYS_FILES()
    return print( "Clear" )

if __name__ == "__main__":
  file_path = arg_obtion.data_path
  save_path = arg_obtion.save_path
  file_name = arg_obtion.file_name
  main(file_path, save_path, file_name)
