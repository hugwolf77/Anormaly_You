
import os
import json
import pandas as pd
from typing import Set, List, Dict, Tuple
from openpyxl import Workbook, load_workbook

from dataform import TS_DATA


class LOAD_TS_FILE:
    def __init__(self, file_path, file_format, meta_file_format, file_ix, target_ix) -> None:
        self.file_ix = file_ix
        self.target_ix = target_ix
        self.file_path = file_path
        self.file_format = file_format
        self.meta_file_format = meta_file_format

        self.file_list = [f"{file}" for file in os.listdir(self.file_path+'csv/') if self.file_format in file]
        self.file_list.sort()
        self.meta_file_list = [f"{file}" for file in os.listdir(self.file_path+'json/') if self.meta_file_format in file]
        self.meta_file_list.sort()
        self.fileName = self.file_list[self.file_ix]
        self.Metafile_name =  self.meta_file_list[self.target_ix]
        
    def File_loader(self):
        if self.file_format == ".csv":
            tsFile = pd.read_csv(os.path.join(self.file_path+'csv/',self.fileName), header=0)
            print("fime_name : ",self.fileName ," file's shape : ", tsFile.shape)
        else:
            tsFile = pd.read_excel(os.path.join(self.file_path+self.fileName), header=0)
            print("fime_name : ",self.fileName ," file's shape : ", tsFile.shape)
        return tsFile

    def Meta_file_loader(self):
      if self.meta_file_format == ".json":
        with open(os.path.join(self.file_path+'json/',self.Metafile_name)) as meta:
          metainfo = json.load(meta)
          tsMeta = metainfo
          df_input = str(tsMeta['input_serial'])
          df_target = str(tsMeta['output_serial'])
          in_cols = df_input.replace(' ','').split(',')
          in_cols_len = len(in_cols)
          tg_cols = df_target.replace(' ','').split(',')
          tg_cols_len = len(tg_cols)
          print(f"in_cols : {in_cols}, tg_cols : {tg_cols}")
      return in_cols, in_cols_len, tg_cols, tg_cols_len, tsMeta

class REPORT:
    def __init__(self, TS_DATA_ANLYS_LIST: List[TS_DATA], file_path:str, save_path:str, fxls_name:str, sheet_name:str, sheet_col:List[str]) -> None:
        self.TS_DATA_ANLYS_LIST = TS_DATA_ANLYS_LIST
        self.file_path = file_path
        self.save_path = save_path
        self.fxls_name = fxls_name
        self.sheet_name = sheet_name
        self.sheet_col = sheet_col

    def Set_wb(self):
        if self.fxls_name in os.listdir(self.save_path):
            wb = load_workbook(filename = os.path.join(self.save_path+self.fxls_name))
        else:
            wb = Workbook()
        write_ws = wb.create_sheet(self.sheet_name,-1)
        fig_ws = wb.create_sheet(self.sheet_name+'_plot',-1)
        write_ws = wb.active
        colname = ['순번','measure_id','reservoir','name','target_value','input_col','modify_day','file_name','file_path','row_count', 'MAE', 'MSE','RMSE','MAPE','MSPE','RSE','CORR'] 
        colname.extend([])
        write_ws.append(colname)
        for TSDATA in self.TS_DATA_ANLYS_LIST:
            row_max = write_ws.max_row + 1
            col_max = write_ws.max_column # need sheet col lenth check
            write_ws.cell(row_max,1,TSDATA.Id)
            write_ws.cell(row_max,2,TSDATA.ob_measure_id)
            write_ws.cell(row_max,3,TSDATA.ob_reservoir)
            write_ws.cell(row_max,4,TSDATA.ob_name)
            write_ws.cell(row_max,5,TSDATA.ob_target)
            write_ws.cell(row_max,6,TSDATA.ob_input)
            write_ws.cell(row_max,7,TSDATA.Modi_Date)
            write_ws.cell(row_max,8,TSDATA.FileName)
            write_ws.cell(row_max,9,TSDATA.FilePath)
            write_ws.cell(row_max,10,TSDATA.Row_count)
            write_ws.cell(row_max,11,TSDATA.MAE)
            write_ws.cell(row_max,12,TSDATA.MSE)
            write_ws.cell(row_max,13,TSDATA.RMSE)
            write_ws.cell(row_max,14,TSDATA.MAPE)
            write_ws.cell(row_max,15,TSDATA.MSPE)
            write_ws.cell(row_max,16,TSDATA.RSE)
            write_ws.cell(row_max,17,str(TSDATA.CORR))
        wb.save(self.save_path+self.fxls_name)